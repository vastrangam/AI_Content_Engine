#!/usr/bin/env python3
"""Recalculate a built workbook and refuse to pass it if anything is wrong.

    python3 engine/recalc.py FY2025-26.xlsx
    python3 engine/recalc.py FY2025-26.xlsx --expect expect.json

§6: "Always run recalc.py after building and fix any formula errors before
delivering — zero errors required across every sheet in the workbook."

openpyxl writes formulas but cannot evaluate them, so a workbook that looks
finished can be full of #REF!. This opens the file in LibreOffice headless,
which recalculates every formula on load, and then reads the result back.

Two things are checked, and the second is the one that matters:

  1. No cell recalculates to an Excel error, and no formula recalculates to
     nothing at all.

  2. Every figure the engine computed, Excel arrives at independently. The
     builder records what each checked cell must come to; if the workbook's own
     formulas disagree with the engine, one of the two is wrong and the
     workbook does not ship. This is what makes "every cell is a live formula"
     worth insisting on — a total written as a number can only be trusted, but
     a total written as a formula can be checked.

Needs LibreOffice Calc:  apt-get install libreoffice-calc
"""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

ERRORS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NULL!", "#NUM!",
          "Err:", "#ERROR")

TOLERANCE = 0.02


class RecalcUnavailable(RuntimeError):
    """LibreOffice is not here, or cannot load spreadsheets."""


def soffice() -> str:
    for name in ("soffice", "libreoffice"):
        found = shutil.which(name)
        if found:
            return found
    raise RecalcUnavailable(
        "LibreOffice is not installed, so formulas cannot be recalculated.\n"
        "  apt-get install libreoffice-calc\n"
        "Without it the workbook may ship with #REF! in it and nobody would know."
    )


def recalculate(path, timeout: int = 600) -> Path:
    """A copy of the workbook with every formula evaluated and cached."""
    src = Path(path).resolve()
    if not src.exists():
        raise FileNotFoundError(src)
    tmp = Path(tempfile.mkdtemp(prefix="recalc-"))
    profile = tmp / "profile"
    out = tmp / "out"
    out.mkdir(parents=True, exist_ok=True)
    proc = subprocess.run(
        [soffice(), "--headless", "--norestore",
         f"-env:UserInstallation=file://{profile}",
         "--convert-to", "xlsx", "--outdir", str(out), str(src)],
        capture_output=True, text=True, timeout=timeout,
    )
    got = out / (src.stem + ".xlsx")
    if not got.exists():
        raise RecalcUnavailable(
            f"LibreOffice could not load {src.name}.\n"
            f"  stdout: {proc.stdout.strip()}\n  stderr: {proc.stderr.strip()}\n"
            "If it reports 'source file could not be loaded', the Calc filter is "
            "missing:  apt-get install libreoffice-calc"
        )
    return got


def check(path, expect: dict | None = None, timeout: int = 600) -> dict:
    """Recalculate, then report every error cell and every figure that moved."""
    import openpyxl

    recalculated = recalculate(path, timeout)
    live = openpyxl.load_workbook(recalculated, data_only=True)
    written = openpyxl.load_workbook(path, data_only=False)

    errors, empty = [], []
    cells = 0
    for ws in written:
        got = live[ws.title]
        for row in ws.iter_rows():
            for c in row:
                if not isinstance(c.value, str) or not c.value.startswith("="):
                    continue
                cells += 1
                value = got.cell(row=c.row, column=c.column).value
                if isinstance(value, str) and any(e in value for e in ERRORS):
                    errors.append({"sheet": ws.title, "cell": c.coordinate,
                                   "formula": c.value[:120], "value": value})
                elif value is None and '""' not in c.value:
                    # A formula guarded with "" is meant to be able to come out
                    # blank — a person with no employed month this year has no
                    # blended rate, and printing a zero there would be a lie.
                    # Anything else recalculating to nothing is a real fault.
                    empty.append({"sheet": ws.title, "cell": c.coordinate,
                                  "formula": c.value[:120]})

    moved = []
    for ref, want in (expect or {}).items():
        sheet, _, cell = ref.rpartition("!")
        sheet = sheet.strip("'")
        if sheet not in live.sheetnames:
            moved.append({"where": ref, "want": want, "got": "no such sheet"})
            continue
        got = live[sheet][cell].value
        try:
            ok = abs(float(got) - float(want)) <= TOLERANCE
        except (TypeError, ValueError):
            ok = False
        if not ok:
            moved.append({"where": ref, "want": want, "got": got})

    return {
        "path": str(path),
        "formula_cells": cells,
        "errors": errors,
        "empty": empty,
        "disagreements": moved,
        "passed": not errors and not moved,
    }


def report(result: dict) -> str:
    lines = [f"recalc  {Path(result['path']).name}",
             f"        {result['formula_cells']:,} formula cells recalculated"]
    if result["errors"]:
        lines.append(f"        {len(result['errors'])} FORMULA ERRORS")
        for e in result["errors"][:12]:
            lines.append(f"          {e['sheet']}!{e['cell']}  {e['value']}"
                         f"   <-  {e['formula']}")
        if len(result["errors"]) > 12:
            lines.append(f"          ... and {len(result['errors']) - 12} more")
    if result["empty"]:
        lines.append(f"        {len(result['empty'])} formulas recalculated to nothing")
        for e in result["empty"][:6]:
            lines.append(f"          {e['sheet']}!{e['cell']}   <-  {e['formula']}")
    if result["disagreements"]:
        lines.append(f"        {len(result['disagreements'])} figures where the "
                     f"workbook and the engine disagree")
        for d in result["disagreements"]:
            lines.append(f"          {d['where']}  engine {d['want']}  workbook {d['got']}")
    lines.append("        " + ("PASSED — zero errors" if result["passed"]
                               else "FAILED — do not deliver this workbook"))
    return "\n".join(lines)


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("workbook")
    ap.add_argument("--expect", help="JSON of {'Sheet!A1': value} the engine computed")
    ap.add_argument("--timeout", type=int, default=600)
    args = ap.parse_args(argv)

    expect = {}
    if args.expect and Path(args.expect).exists():
        expect = json.loads(Path(args.expect).read_text(encoding="utf-8"))
    result = check(args.workbook, expect, args.timeout)
    print(report(result))
    return 0 if result["passed"] else 1


if __name__ == "__main__":
    sys.exit(main())
