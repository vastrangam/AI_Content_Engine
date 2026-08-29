#!/usr/bin/env python3
"""Write master.json back out as a Staff & Karigar Master Data workbook.

The JSON fixture already reproduces the FY2025-26 figures. This turns it into
the workbook shape from §1.1, so the same data can be maintained in Excel
instead — and so a filled example exists to compare a hand-filled one against.

    python3 engine/tests/fixture_to_template.py Master_Data.xlsx

The personal and banking columns are written empty, on purpose. The fixture has
never held them and this script will not invent them.
"""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "tests"))

from make_template import (BODY, HOURS_HEADERS, KARIGAR_HEADERS, LOG_HEADERS,
                           READ_ME, STAFF_HEADERS, sheet, write)
from openpyxl import Workbook
from openpyxl.styles import Font

from vastrangam import Master
from vastrangam.attendance import SUNDAY
from vastrangam.master import DAILY_WAGE, PIECE_RATE

FIXTURE = ROOT / "fixtures" / "master.json"


def open_row(log, ident):
    """The value in force now — what the base template columns hold (§4.6)."""
    rows = log.rows(ident)
    return rows[-1].value if rows else None


def _category(group: str) -> str:
    """'M' and 'F' in the fixture are 'Male' and 'Female' in the workbook."""
    key = str(group).strip().upper()
    return {"M": "Male", "F": "Female"}.get(key, str(group))


def build(out_path, fixture=FIXTURE):
    m = Master.from_json(fixture)
    wb = Workbook()

    staff = sheet(wb, "Staff Master", STAFF_HEADERS,
                  {"Name": 20, "Address": 30, "Role": 16, "KRA": 22})
    hours = sheet(wb, "Hours Reference", HOURS_HEADERS, {"Notes": 34})
    karigar = sheet(wb, "Karigar Master", KARIGAR_HEADERS, {"Name": 22})
    salary_log = sheet(wb, "Salary Log", LOG_HEADERS)
    hours_log = sheet(wb, "Threshold Hours Log", LOG_HEADERS)
    days_log = sheet(wb, "Threshold Days Log", LOG_HEADERS)
    # THE PAY BASIS NEEDS ITS OWN HISTORY, not just the value in force.
    # Staff Master carries one basis per person, which is the basis TODAY. Somebody who
    # was hourly one year and on piece rate the next has two, and writing only the second
    # backdated the new basis over the whole of the old year — the workbook then priced
    # last year at this year's rules and nothing looked wrong.
    basis_log = sheet(wb, "Pay Basis Log", LOG_HEADERS)

    staff_rows = []
    for ident in sorted(m.people):
        p = m.people[ident]
        spells = m.employment.spells(ident)
        basis = open_row(m.pay_basis, ident)
        staff_rows.append((
            ident, p.name,
            "Male" if p.gender.upper().startswith("M") else "Female",
            ", ".join(p.roles) or None, None,
            None, None, None, None, None, None, None, None,   # personal: left empty
            open_row(m.salary, ident),
            open_row(m.threshold_hours, ident),
            open_row(m.threshold_days, ident),
            open_row(m.daily_wage, ident),
            spells[0].joined if spells else None,
            spells[-1].left if spells and spells[-1].left else None,
            p.roster,
            # Written out explicitly rather than left to §1.2's inference, which
            # would read flat pay as attendance-based wherever both a salary and
            # a threshold are filled.
            basis,
        ))
    write(staff, staff_rows)

    # The category must be written with the same word the Gender column uses,
    # or nothing matches and every month goes unresolvable.
    # A person with their own clock is written against their NAME in the same column —
    # the loader asks the alias table which kind of row it is reading.
    write(hours, [
        (_category(group), "Sunday / Weekly Off" if kind == SUNDAY else kind,
         value, None)
        for (group, kind), value in sorted(m.shift_hours.items())
    ] + [
        (m.people[ident].name if ident in m.people else ident,
         "Sunday / Weekly Off" if kind == SUNDAY else kind, value,
         "this person's own clock, not their category's")
        for (ident, kind), value in sorted(m.shift_hours_by_person.items())
    ])

    write(karigar, [(None,) * len(KARIGAR_HEADERS)][:0] or [])

    def history(log, ws):
        rows = []
        for ident in sorted(m.people):
            got = log.rows(ident)
            if len(got) > 1:                       # one row is the current value
                for r in got:
                    rows.append((m.people[ident].name, r.frm, r.to, r.value))
        write(ws, rows)

    history(m.salary, salary_log)
    history(m.threshold_hours, hours_log)
    history(m.threshold_days, days_log)
    # Every row, not only where there is more than one: Staff Master's Pay Basis column
    # is dated from the joining date, so a single row that starts later is exactly the
    # case this tab exists to correct.
    write(basis_log, [(m.people[ident].name, r.frm, r.to, r.value)
                      for ident in sorted(m.people)
                      for r in m.pay_basis.rows(ident)])

    def tab(title, headers, rows):
        ws = wb.create_sheet(title)
        ws.sheet_view.showGridLines = False
        for i, h in enumerate(headers, 1):
            c = ws.cell(1, i, h)
            c.fill, c.font = staff.cell(1, 1).fill.copy(), staff.cell(1, 1).font.copy()
        write(ws, rows)
        return ws

    # A PERSON'S RATE PER HOUR, AND SEPARATELY THE OPERATION'S RATE PER PIECE.
    # These were one tab keyed by Staff with a Unit column saying which. That put two
    # different facts in one place: the hourly figure is genuinely somebody's, and the
    # per-piece figure is the operation's and shared by everyone doing that work. A
    # workbook that cannot tell them apart cannot add the fourth person to an operation
    # without also inventing a rate for them.
    tab("Hourly Rate Log", ["Staff", "Operation", "Effective From", "Effective To", "Rate",
                            "Unit"],
        [(m.people[ident].name, (r.value or {}).get("operation")
          if isinstance(r.value, dict) else None, r.frm, r.to,
          r.value.get("rate") if isinstance(r.value, dict) else r.value, "per hour")
         for ident in sorted(m.people) for r in m.hourly_rate.rows(ident)])

    tab("Piece Rate Card", ["Operation", "Garment", "Effective From", "Effective To", "Rate"],
        [tuple(str(r.key).split("|", 1)) + (r.frm, r.to, r.value)
         for r in sorted(m.piece_rate.rows(), key=lambda r: (str(r.key), r.frm))
         if "|" in str(r.key)])

    notes = wb.create_sheet("Read Me", 0)
    notes.sheet_view.showGridLines = False
    text = READ_ME.strip() + "\n\n" + FILLED_NOTE.strip()
    for i, line in enumerate(text.splitlines(), 1):
        c = notes.cell(i, 1, line)
        c.font = Font(name="Arial", size=11, bold=line.endswith(":") or i == 1)
    notes.column_dimensions["A"].width = 100

    wb.save(out_path)
    return out_path


FILLED_NOTE = """
About this filled copy:
  The salaries, thresholds, joining dates and pay bases here are the ones that
  reproduce the FY2025-26 report — 9,75,648.80 payroll against the 9,75,649 in
  your own file. Check them against what you know and correct anything wrong.
  The personal and banking columns are deliberately empty. Fill them in only in
  the copy you keep with your payroll records, and never in one that goes into a
  code repository or an email.
  Pay Basis is written out here rather than left blank, because the inference in
  §1.2 reads flat pay as attendance-based whenever a salary and a threshold are
  both filled.
"""


if __name__ == "__main__":
    out = Path(sys.argv[1] if len(sys.argv) > 1 else "Master_Data_filled.xlsx")
    build(out)
    print(f"wrote {out}")
