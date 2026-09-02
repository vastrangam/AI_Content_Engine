#!/usr/bin/env python3
"""Build a Staff & Karigar Master Data workbook to §1.1's exact shape.

Two uses. It is the blank template a company fills in, and — filled with
deliberately awkward rows — it is what the reader is tested against until the
real one arrives.

    python3 engine/tests/make_template.py blank.xlsx          # empty, to fill in
    python3 engine/tests/make_template.py fixture.xlsx --demo # filled, for tests

The demo rows are invented. They exercise every branch of §1.2's pay-basis
inference, including the combination the prompt cannot price.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADER_FILL = PatternFill("solid", fgColor="1F3B4D")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
FILL_IN = PatternFill("solid", fgColor="FFF6CC")     # yellow: yours to type in
BODY = Font(name="Arial", size=10)

STAFF_HEADERS = [
    "Staff ID", "Name", "Gender", "Role", "KRA", "Phone", "Address",
    "Aadhaar No.", "PAN No.", "Bank Name", "Account No.", "IFSC", "UPI ID",
    "Monthly Salary (₹)", "Threshold Hour", "Threshold Day", "Daily Wage (₹)",
    "Join Date", "Leave Date", "Status", "Pay Basis",
]

HOURS_HEADERS = ["Category", "Day Type", "Present (P) Hours", "Notes"]

KARIGAR_HEADERS = [
    "Karigar ID", "Name", "KRA", "Phone", "Address", "Aadhaar No.", "PAN No.",
    "Bank Name", "Account No.", "IFSC", "UPI ID", "Earning (Per Piece Wise)",
    "Join Date", "Status",
]

LOG_HEADERS = ["Staff", "Effective From", "Effective To", "Rate"]

# Invented. Not anyone's real data — the reader is what is under test here.
DEMO_STAFF = [
    # id     name        g   role       salary  thr_h thr_d wage  join        status   basis
    ("S001", "Aarav",   "Male",   "Cutting", 45000, 280,  28,   None, "01-08-2025", "Active",   None),
    ("S002", "Bhavna",  "Female", "Packing",  9000, 230,  28,   None, "01-04-2025", "Active",   None),
    ("S003", "Chetan",  "Male",   "QC",      18000, 280,  28,   None, "01-04-2025", "Active",   "Flat"),
    ("S004", "Divya",   "Female", "Helper",   None, None, None,  450, "01-06-2025", "Active",   None),
    ("S005", "Eshan",   "Male",   "Iron",     None, None, None, None, "01-04-2025", "Active",   None),
    ("S006", "Farhan",  "Male",   "Dispatch", None, 280,  None, None, "01-04-2025", "Active",   None),
    ("S007", "Gita",    "Female", "Sampling", 12000, None, None, None, "01-04-2025", "Inactive", None),
]

# Written the way the company's own template writes it — full words for the
# category, and its own wording for the rest day. The reader canonicalises both.
DEMO_HOURS = [
    ("Male", "Weekday", 10.0, "09:30-20:00 less 30 min lunch"),
    ("Male", "Sunday / Weekly Off", 5.0, "09:30-15:00 less 30 min lunch"),
    ("Female", "Weekday", 8.0, "09:30-18:00 less 30 min lunch"),
    ("Female", "Sunday / Weekly Off", 5.5, "09:30-15:30 less 30 min lunch"),
]

DEMO_KARIGAR = [
    ("K001", "Unit One", "Stitching", 120.0, "01-04-2025", "Active"),
    ("K002", "Unit Two & team", "Stitching", 135.0, "01-04-2026", "Active"),
]

DEMO_SALARY_LOG = [
    ("Chetan", "01-04-2025", "31-05-2025", 15000),
    ("Chetan", "01-06-2025", None, 18000),
]

DEMO_THR_HOURS_LOG = [
    ("Aarav", "01-08-2025", "31-10-2025", 280),
    ("Aarav", "01-11-2025", None, 270),
]

DEMO_THR_DAYS_LOG = [
    ("Aarav", "01-08-2025", "31-10-2025", 28),
    ("Aarav", "01-11-2025", None, 27),
]


def sheet(wb, title, headers, widths=None):
    ws = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] else wb.active
    ws.title = title
    ws.sheet_view.showGridLines = False
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = (widths or {}).get(h, 16)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    return ws


def write(ws, rows, start=2):
    for r, values in enumerate(rows, start):
        for c, v in enumerate(values, 1):
            cell = ws.cell(r, c, v)
            cell.font = BODY
    return start + len(rows)


def build(path, demo=False, blank_rows=40):
    wb = Workbook()

    staff = sheet(wb, "Staff Master", STAFF_HEADERS,
                  {"Name": 20, "Address": 30, "Role": 16, "KRA": 22})
    hours = sheet(wb, "Hours Reference", HOURS_HEADERS, {"Notes": 34})
    karigar = sheet(wb, "Karigar Master", KARIGAR_HEADERS, {"Name": 22, "Address": 30})
    salary_log = sheet(wb, "Salary Log", LOG_HEADERS)
    hours_log = sheet(wb, "Threshold Hours Log", LOG_HEADERS)
    days_log = sheet(wb, "Threshold Days Log", LOG_HEADERS)

    if demo:
        write(staff, [
            (sid, name, g, role, f"{role} owner", None, None, None, None, None,
             None, None, None, sal, thr_h, thr_d, wage, join, None, status, basis)
            for sid, name, g, role, sal, thr_h, thr_d, wage, join, status, basis in DEMO_STAFF
        ])
        write(hours, DEMO_HOURS)
        write(karigar, [
            (kid, name, kra, None, None, None, None, None, None, None, None,
             rate, join, status)
            for kid, name, kra, rate, join, status in DEMO_KARIGAR
        ])
        write(salary_log, DEMO_SALARY_LOG)
        write(hours_log, DEMO_THR_HOURS_LOG)
        write(days_log, DEMO_THR_DAYS_LOG)
        last = max(len(DEMO_STAFF) + 1, 2)
    else:
        last = blank_rows + 1
        for ws in (staff, hours, karigar, salary_log, hours_log, days_log):
            for r in range(2, blank_rows + 2):
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = FILL_IN
                    ws.cell(r, c).font = BODY

    # Dropdowns, so a status or a basis can never be a typo.
    _validate(staff, "T", f'"Active,Inactive"', last if demo else blank_rows + 1,
              STAFF_HEADERS.index("Status") + 1,
              "Status must be Active or Inactive")
    _validate(staff, "U", '"Flat,Attendance,Daily-wage,Piece-rate"',
              last if demo else blank_rows + 1,
              STAFF_HEADERS.index("Pay Basis") + 1,
              "Leave blank to infer the basis from which columns are filled (§1.2), "
              "or state it here to override the inference")
    _validate(hours, "B", '"Weekday,Sunday / Weekly Off"',
              last if demo else blank_rows + 1, 2,
              "Add one row per Category and Day Type whose hours actually differ")

    notes = wb.create_sheet("Read Me", 0)
    notes.sheet_view.showGridLines = False
    for i, line in enumerate(READ_ME.strip().splitlines(), 1):
        c = notes.cell(i, 1, line)
        c.font = Font(name="Arial", size=11, bold=line.endswith(":") or i == 1)
    notes.column_dimensions["A"].width = 100

    wb.save(path)
    return path


def _validate(ws, letter, formula, last_row, col, prompt):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                        showErrorMessage=True)
    dv.error = prompt
    dv.errorTitle = "Not one of the allowed values"
    dv.prompt, dv.promptTitle = prompt, ws.cell(1, col).value
    ws.add_data_validation(dv)
    dv.add(f"{letter}2:{letter}{max(last_row, 200)}")


READ_ME = """
Staff & Karigar Master Data

This workbook is the only place your company's data lives. The engine holds the
rules; this holds the people, the rates and the hours. Fill it in, and the
reports follow.

Staff Master:
  One row per person. Yellow cells are yours to fill.
  Pay Basis can be left blank — it is worked out from which columns you filled:
    Monthly Salary + Threshold Hour   -> Attendance
    Daily Wage, no salary             -> Daily-wage
    Monthly Salary only               -> Flat
    none of the three                 -> Piece-rate (rate lives in the work report)
  Fill Pay Basis in only when you want to override that.
  Threshold Day matters: it is not the same as Threshold Hour divided by a day.
  280 hours at a ten-hour day is 28 days, but 230 hours at an eight-hour day is
  28.75, and the real threshold is 28. Type it; do not let it be guessed.
  Status Inactive keeps the history without putting the person on the current
  roster. Never delete a row — an ex-employee still has last year's pay.

Hours Reference:
  One row per Category and Day Type whose hours genuinely differ. If everyone
  works the same hours, one row with Category and Day Type left blank is enough.
  A category with no row here is an error, not zero hours.

Karigar Master:
  Identity, contact and payment details, plus a headline per-piece figure.
  The real per-design rate card is a separate file — this number is a reference.

Salary Log / Threshold Hours Log / Threshold Days Log:
  Optional. Use them when a rate changed part-way through and you want past
  months priced at what applied then. Close the old row with an Effective To,
  open a new one from the date of the change, and never delete history.

A note on the personal columns:
  Aadhaar, PAN, bank account, IFSC and UPI are read only when a payment run
  needs them. They are never copied into any report the engine produces, and
  this workbook should never be committed to a code repository or emailed
  around. Keep it where your payroll records already live.
"""


if __name__ == "__main__":
    out = Path(sys.argv[1] if len(sys.argv) > 1 else "master_template.xlsx")
    demo = "--demo" in sys.argv
    build(out, demo)
    print(f"wrote {out}{' (with demo rows)' if demo else ' (blank)'}")
