"""How every sheet in the deliverable looks — Combined Master Prompt §6.

    Dark slate header rows (white bold text), purple/pastel accent for subtotal
    rows, gridlines off, frozen header panes, Arial font, currency format with
    2 decimals, blue text for editable input cells.

One place, so a sheet cannot quietly look different from its neighbours, and so
a change of house style is a change here rather than in nineteen builders.

openpyxl is imported lazily for the same reason it is in xlsx.py: the rules and
their self-tests must run on a machine that has never installed it.
"""

from __future__ import annotations

FONT = "Arial"

SLATE = "FF334155"          # header rows
SLATE_LIGHT = "FF475569"    # a second header band, where a sheet needs two
ACCENT = "FFEDE9FE"         # subtotal rows — the purple/pastel accent
ACCENT_EDGE = "FFC4B5FD"
INPUT_BLUE = "FF1D4ED8"     # editable input cells, and only those
MUTED = "FF64748B"
WARN = "FFFEF3C7"

MONEY = '#,##0.00'
QTY = '#,##0.00'
INT = '#,##0'
PERCENT = '0.0%'
DATE = 'dd-mmm-yyyy'
RATE = '#,##0.0000'

# Excel has no "never" date. A far-future sentinel is what makes an open-ended
# effective-dated row comparable with <= and >= like every closed one, which is
# what §3.6.1 asks the Effective To (calc) column to provide.
FOREVER = "DATE(2999,12,31)"


def _mods():
    try:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:  # pragma: no cover - depends on the machine
        raise RuntimeError(
            "openpyxl is needed to write workbooks.  pip install openpyxl"
        ) from exc
    return Alignment, Border, Font, PatternFill, Side


class Style:
    """The house style, resolved once per workbook."""

    def __init__(self):
        Alignment, Border, Font, PatternFill, Side = _mods()
        self.header_font = Font(name=FONT, bold=True, color="FFFFFFFF", size=10)
        self.title_font = Font(name=FONT, bold=True, size=13, color=SLATE)
        self.note_font = Font(name=FONT, size=9, color=MUTED, italic=True)
        self.body_font = Font(name=FONT, size=10)
        self.bold_font = Font(name=FONT, size=10, bold=True)
        self.input_font = Font(name=FONT, size=10, color=INPUT_BLUE, bold=True)
        self.header_fill = PatternFill("solid", fgColor=SLATE)
        self.header_fill_2 = PatternFill("solid", fgColor=SLATE_LIGHT)
        self.accent_fill = PatternFill("solid", fgColor=ACCENT)
        self.warn_fill = PatternFill("solid", fgColor=WARN)
        self.wrap = Alignment(vertical="top", wrap_text=True)
        self.centre = Alignment(horizontal="center", vertical="center")
        self.top = Alignment(vertical="top")
        thin = Side(style="thin", color=ACCENT_EDGE)
        self.accent_border = Border(top=thin, bottom=thin)


def prepare(ws, style: Style) -> None:
    """Gridlines off, and Arial everywhere by default."""
    ws.sheet_view.showGridLines = False


def title(ws, style: Style, row: int, text: str, note: str = "") -> int:
    """A sheet title, and the sentence under it saying what the sheet is for.

    Returns the next free row, so a caller never counts rows by hand.
    """
    c = ws.cell(row=row, column=1, value=text)
    c.font = style.title_font
    row += 1
    if note:
        n = ws.cell(row=row, column=1, value=note)
        n.font = style.note_font
        row += 1
    return row + 1


def header(ws, style: Style, row: int, labels, widths=None, band=1) -> int:
    """One header row, frozen by the caller. Returns the first data row."""
    fill = style.header_fill if band == 1 else style.header_fill_2
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = style.header_font
        c.fill = fill
        c.alignment = style.wrap
    if widths:
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, start=1):
            if w:
                ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    return row + 1


def write_row(ws, style: Style, row: int, values, formats=None,
              subtotal=False, bold=False) -> int:
    """One row of values. A value beginning with '=' is written as a formula.

    Returns the next row, for the same reason header() does.
    """
    formats = formats or []
    for i, value in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=value)
        c.font = style.bold_font if (subtotal or bold) else style.body_font
        if subtotal:
            c.fill = style.accent_fill
            c.border = style.accent_border
        fmt = formats[i - 1] if i - 1 < len(formats) else None
        if fmt:
            c.number_format = fmt
    return row + 1


def note(ws, style: Style, row: int, text: str) -> int:
    c = ws.cell(row=row, column=1, value=text)
    c.font = style.note_font
    return row + 1


def paragraphs(ws, style: Style, row: int, lines) -> int:
    """The Read Me sheets. Plain sentences, one per row, blanks preserved."""
    for line in lines:
        if line.startswith("## "):
            c = ws.cell(row=row, column=1, value=line[3:])
            c.font = style.bold_font
        else:
            c = ws.cell(row=row, column=1, value=line)
            c.font = style.body_font
            c.alignment = style.top
        row += 1
    return row


def editable(ws, style: Style, row: int, col: int, value, fmt=None):
    """§6 — blue text for editable input cells, and nowhere else."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = style.input_font
    if fmt:
        c.number_format = fmt
    return c


def total(column: str, first: int, last: int) -> str:
    """A SUM over a block that may be empty.

    A sheet with no data rows would otherwise write SUM(B6:B5) — a backwards
    range that Excel reads as #VALUE!. An FY with no work report is a normal
    thing to have, so its total is a plain zero rather than an error.
    """
    return f"=SUM({column}{first}:{column}{last})" if last >= first else 0


def col(index: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(index)


def quote(sheet_name: str) -> str:
    """A sheet reference that survives a space in the name."""
    return f"'{sheet_name}'" if any(ch in sheet_name for ch in " -&()") else sheet_name


def text_arg(value) -> str:
    """A string safely embedded in a formula."""
    return '"' + str(value).replace('"', '""') + '"'
