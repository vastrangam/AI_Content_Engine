"""The deliverable — one workbook per financial year.

Sheets, in the order §4 asks for: the two Read Me sheets, then the Combined
Productivity Overview, then the karigar pipeline's nine sheets (§2.4) and the
staff pipeline's nine (§3.6).

The governing rule is §4's, and it applies to the whole workbook rather than
only to the Overview:

    Every cell is a live formula referencing the two pipelines' own sheets —
    never a typed-in number.

So a total is never written as a number the engine computed. It is written as
the formula that computes it from the rows beside it, and recalc.py then opens
the workbook, recalculates every formula, and checks that what Excel arrives at
is what the engine arrived at. A figure that only the engine can produce is a
figure nobody can audit.

What IS written as a value: the source facts. An attendance mark, a salary from
the log, a production quantity, a rate. Everything derived from them is a
formula.
"""

from __future__ import annotations

import datetime as dt
from collections import defaultdict
from dataclasses import dataclass, field

from . import sheetstyle as S
from .calendar_util import Month, fy_months
from .karigar import BOTTOM, DUPATTA, TOP
from .master import ATTENDANCE, DAILY_WAGE, FLAT, PIECE_RATE
from .pay import EMPLOYED, NOT_EMPLOYED, NO_DATA, UNRESOLVED, blended_hourly
from .logs import Unresolved

READ_ME_KARIGAR = "Read Me - Karigar"
READ_ME_STAFF = "Read Me - Staff"
OVERVIEW = "Combined Productivity Overview"

K_ITEMWISE = "Item-wise Production & Cost"
K_COMBINED = "Combined Production"
K_EARNINGS = "Karigar Earnings"
K_DETAIL = "Karigar x Design Detail"
K_PAYSUMMARY = "Payment Summary (Yearly)"
K_PAYDETAIL = "Payment Detail (Monthly)"
K_RAW = "Raw Transactions (Audit)"
K_REVIEW = "Needs Review"

ST_SALARY = "Salary Log"
ST_THRESHOLD = "Threshold Log"
ST_MASTER = "Staff Master"
ST_ATTENDANCE = "Attendance"
ST_MONTHLY = "Monthly Summary"
ST_PAYDETAIL = "Payment Detail"
ST_RECON = "Payment Reconciliation"
ST_WORK = "Work Report - Cost per Design"
ST_PERFORMANCE = "Performance Tracking"

SHEET_ORDER = [
    READ_ME_KARIGAR, READ_ME_STAFF, OVERVIEW,
    K_ITEMWISE, K_COMBINED, K_EARNINGS, K_DETAIL, K_PAYSUMMARY, K_PAYDETAIL,
    K_RAW, K_REVIEW,
    ST_SALARY, ST_THRESHOLD, ST_MASTER, ST_ATTENDANCE, ST_MONTHLY,
    ST_PAYDETAIL, ST_RECON, ST_WORK, ST_PERFORMANCE,
]

FY_TOTAL = "FY Total"
NO_MONTH_BREAKDOWN = "whole FY (no month breakdown)"


@dataclass
class Inputs:
    """Everything one FY's workbook is built from.

    A missing side is not an error. §0 is explicit: build whatever was provided,
    skip the other pipeline's sheets, and say so plainly rather than fabricating
    the missing half.
    """

    fy: str
    master: object = None
    book: object = None
    payroll: dict = field(default_factory=dict)
    allocation: dict = field(default_factory=dict)
    work_rows: list = field(default_factory=list)
    quantities: dict = field(default_factory=dict)
    payments: dict = field(default_factory=dict)
    karigar: object = None
    review: list = field(default_factory=list)

    @property
    def has_staff(self) -> bool:
        return self.master is not None and bool(self.payroll)

    @property
    def has_karigar(self) -> bool:
        return self.karigar is not None


@dataclass
class Built:
    path: str
    sheets: list = field(default_factory=list)
    skipped: list = field(default_factory=list)
    # What the engine says each checked cell should come to. recalc.py compares
    # the recalculated workbook against this, which is the whole point of
    # writing formulas instead of numbers.
    expect: dict = field(default_factory=dict)
    notes: list = field(default_factory=list)


class _Build:
    def __init__(self, inputs: Inputs):
        import openpyxl

        self.inp = inputs
        self.fy = inputs.fy
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)
        self.style = S.Style()
        self.out = Built(path="")
        self.months = list(fy_months(self.fy))

        # Layout facts other sheets need. Filled in as each sheet is written,
        # never guessed — a formula that points at the wrong row is worse than
        # no formula at all.
        self.att_cols: dict[str, tuple[str, str]] = {}   # staff -> (status, hours)
        self.att_rows: tuple[int, int] = (0, 0)
        self.salary_rows: tuple[int, int] = (0, 0)
        self.threshold_rows: tuple[int, int] = (0, 0)
        self.master_rows: tuple[int, int] = (0, 0)
        self.hours_ref_rows: tuple[int, int] = (0, 0)
        self.monthly_rows: tuple[int, int] = (0, 0)
        self.work_staff_rows: dict[str, int] = {}
        self.work_total_cost_cell = ""
        self.recon_total_cells: dict[str, str] = {}
        self.k_totals: dict[str, str] = {}

    # -- helpers ------------------------------------------------------------

    def sheet(self, name):
        return self.wb[name]

    def staff_order(self) -> list:
        m = self.inp.master
        if m is None:
            return []
        return [s for s in sorted(m.people) if m.people[s].status == "OK"]

    def basis(self, staff, month=None):
        try:
            return self.inp.master.basis_of(staff, month or self.months[0])
        except Exception:
            return ATTENDANCE

    def any_basis(self, staff) -> str:
        """The basis in force at any point this year — what the Staff Master
        column should say. A basis that changed mid-year is reported as both."""
        seen = []
        for m in self.months:
            if not self.inp.master.employed(staff, m):
                continue
            try:
                b = self.inp.master.basis_of(staff, m)
            except Exception:
                continue
            if b not in seen:
                seen.append(b)
        return " / ".join(seen) if seen else ATTENDANCE

    def rows_of(self, staff) -> list:
        return [r for r in self.inp.payroll.get("rows", []) if r.staff == staff]

    # =======================================================================
    # THE READ ME SHEETS
    # =======================================================================

    def read_me_karigar(self):
        ws = self.sheet(READ_ME_KARIGAR)
        S.prepare(ws, self.style)
        ws.column_dimensions["A"].width = 118
        row = S.title(ws, self.style, 1, f"Karigar pipeline — methodology, FY{self.fy}")
        k = self.inp.karigar
        lines = [
            "## What this half of the workbook is",
            "Piece-rate stitching production, set-matching, earnings, and payment and "
            "outstanding tracking. It is priced per piece produced, which is why it "
            "shares nothing with the staff half except the Overview sheet.",
            "",
            "## Where the figures come from",
            "Every figure is recomputed from the transaction rows on the "
            f"'{K_RAW}' sheet. Nothing is read off a totals row in the source. The "
            "source's own totals are used only to check the answer, and where they "
            "disagree the disagreement is on the Needs Review sheet rather than "
            "quietly resolved.",
            "",
            "## Set matching (§2.2)",
            "Each design's production is pooled into Top, Bottom and Dupatta slot "
            "totals. A full set counts toward all three; Top & Bottom toward Top and "
            "Bottom only; a Dupatta-only row toward Dupatta only.",
            "",
            "Total Complete Sets is the smallest POPULATED slot. A slot with no "
            "pieces at all drops out of the minimum rather than zeroing the design — "
            "so a design made as Anarkali and Dupatta with no Plazo counts the sets "
            "that were genuinely made and paid for.",
            "",
            "Surplus above the bottleneck is reported, never merged away: bodies "
            "waiting on a dupatta are Pending Dupatta, spare dupattas are Extra "
            "Dupatta, and Top and Bottom extras that do not match each other are "
            "reported separately.",
            "",
            "## Rates (§2.2, §2.3)",
            "The rate on a Complete Sets line is the quantity-weighted average of the "
            "rates actually paid, falling back to the master rate for a design nobody "
            "has been paid for yet. Where rounding leaves a design a rupee off its raw "
            "recorded value, a Rate variance adjustment line carries the difference so "
            "the total ties exactly and the difference stays visible.",
            "",
            "Earned is production value. Paid comes from the source's own Paid column "
            "or payment ledger, matched to the same karigar labels used in production. "
            "Outstanding is Earned less Paid, and it is a live formula everywhere it "
            "appears.",
            "",
            "Job-work vendor rates run higher than in-house rates because they bundle "
            "stitching, cutting, QC, ironing and packing. That is expected and is not "
            "an error.",
            "",
            "## Naming (§2.1)",
            "Each source keeps its own literal naming. Names are never merged across "
            "files, not even across near-matches — a near match is listed on Needs "
            "Review for a person to confirm, and until they do it stays two units.",
        ]
        if k is not None:
            t = k.totals
            lines += [
                "",
                "## This year's parse",
                f"{t['rows']:,} production rows · {t['units']} paying units · "
                f"{t['designs']} designs · {t['pieces']:,.1f} pieces · "
                f"{t['complete_sets']:,} complete sets.",
            ]
        S.paragraphs(ws, self.style, row, lines)

    def read_me_staff(self):
        ws = self.sheet(READ_ME_STAFF)
        S.prepare(ws, self.style)
        ws.column_dimensions["A"].width = 118
        row = S.title(ws, self.style, 1, f"Staff pipeline — methodology, FY{self.fy}")
        lines = [
            "## What this half of the workbook is",
            "Monthly earnings, payment reconciliation and outstanding, cost per piece "
            "by design, and performance tracking — from attendance, work-report task "
            "hours, and payment data.",
            "",
            "Stitching cost is never folded into the Work Report. That sheet is "
            "labour-cost-only by design, because staff overhead and karigar stitching "
            "answer different questions and adding them together answers neither.",
            "",
            "## How a month is priced (§3.5)",
            "Pay is driven by attendance DAYS, not hours.",
            "",
            "Actual Days-Equivalent = Present + Holiday + 0.5 x Half-days. A holiday "
            "pays as a full present day. Absent and blank contribute nothing.",
            "",
            "Daily Rate = the monthly salary in force that month, divided by the "
            "threshold days in force that month. Both are resolved from the "
            "effective-dated logs for that specific month, never from today's value.",
            "",
            "Earning = Daily Rate x Actual Days-Equivalent, uncapped in both "
            "directions: 30 days worked against a 27-day threshold pays for 30.",
            "",
            "A staff member on the Flat basis is paid their full resolved monthly "
            "salary every month regardless of attendance. A staff member on the "
            "Piece-rate basis has no salary, no threshold and no attendance row at "
            "all — their wage is the hours logged against designs in the Work Report "
            "times their flat rate per hour.",
            "",
            "## Hours (§3.5, §3.6.3)",
            "Hours never price a month. They are a reference column, and their only "
            "job is the Work Report's cost per piece, because design task-hours are "
            "logged in hours.",
            "",
            "The reference table is at the top of the Staff Master sheet: a present "
            "day is 10 hours for a man and 8 for a woman on a weekday, 5 and 5.5 on a "
            "Sunday. A half-day is half of that. Absent, holiday and blank are zero.",
            "",
            "The Blended FY Hourly Rate the Work Report costs at is the Blended FY "
            "Daily Rate divided by that person's own weekday shift — not a salary "
            "divided by an hours threshold. Threshold Hrs/Mo survives on the Staff "
            "Master sheet as a legacy reference column and drives nothing.",
            "",
            "The blended rate averages the months a person was actually employed. "
            "Averaging an eight-month spell across twelve months would understate "
            "their rate by a third, and that understatement would then be spread "
            "across every design they touched.",
            "",
            "## Three states, not two (§3.7)",
            "A month is Not Employed, No Data, or a real month. A blank month inside "
            "an employment spell is a tracking gap, not a month of total absence, and "
            "Performance Tracking calls it No Data rather than Below Average.",
            "",
            "## What is flagged rather than assumed",
            "Any fully blank staff-month. Any FY missing payment or work-report data. "
            "Any staff whose status is unconfirmed but whose data appears anyway. Any "
            "staff on Flat or Piece-rate pay, because their earning is deliberately "
            "decoupled from attendance. All of it is on Needs Review.",
        ]
        if self.inp.has_staff:
            p = self.inp.payroll
            states = defaultdict(int)
            for r in p["rows"]:
                states[r.state] += 1
            lines += [
                "",
                "## This year's parse",
                f"{states[EMPLOYED]} employed months · {states[NO_DATA]} no data · "
                f"{states[NOT_EMPLOYED]} not employed · {states[UNRESOLVED]} unresolvable.",
            ]
        S.paragraphs(ws, self.style, row, lines)

    # =======================================================================
    # THE COMBINED PRODUCTIVITY OVERVIEW  (§4)
    # =======================================================================

    def overview(self):
        ws = self.sheet(OVERVIEW)
        S.prepare(ws, self.style)
        row = S.title(
            ws, self.style, 1, f"Karigar & Staff — Combined Productivity Overview, FY{self.fy}",
            "Every cell here is a live formula pointing at the two pipelines' own "
            "sheets. Nothing on this sheet is typed in, and nothing on it can be "
            "edited — it is a rollup, not a data-entry point.")
        row = S.header(ws, self.style, row, ["Metric", "This FY", "Comes from"],
                       [52, 20, 46])

        k, has_k = self.k_totals, self.inp.has_karigar
        has_s = self.inp.has_staff
        blank_k = "No karigar data uploaded for this FY"
        blank_s = "No staff data uploaded for this FY"

        first = row
        def line(label, formula, source, fmt=S.MONEY, present=True, absent=""):
            nonlocal row
            if present:
                row = S.write_row(ws, self.style, row, [label, formula, source],
                                  [None, fmt, None])
            else:
                row = S.write_row(ws, self.style, row, [label, None, absent])

        line("Total Karigar Production Value (stitching earnings)",
             f"={k.get('earned', '')}" if has_k else None,
             f"{K_EARNINGS} grand total", present=has_k, absent=blank_k)
        karigar_earned = f"B{row - 1}" if has_k else None
        line("Total Karigar Payments Made",
             f"={k.get('paid', '')}" if has_k else None,
             f"{K_PAYSUMMARY} grand total", present=has_k, absent=blank_k)
        line("Total Karigar Outstanding",
             f"={k.get('outstanding', '')}" if has_k else None,
             f"{K_PAYSUMMARY} grand total", present=has_k, absent=blank_k)
        karigar_out = f"B{row - 1}" if has_k else None

        staff_earn_src = (f"=SUMIFS({S.quote(ST_MONTHLY)}!$M${self.monthly_rows[0]}:"
                          f"$M${self.monthly_rows[1]},"
                          f"{S.quote(ST_MONTHLY)}!$B${self.monthly_rows[0]}:"
                          f"$B${self.monthly_rows[1]},{S.text_arg('<>' + FY_TOTAL)})")
        line("Total Staff Payroll Earning (all pay bases)",
             staff_earn_src if has_s else None,
             f"{ST_MONTHLY}, every month row", present=has_s, absent=blank_s)
        staff_earned = f"B{row - 1}" if has_s else None
        line("Total Staff Payments Made",
             f"={self.recon_total_cells.get('paid', '')}" if has_s else None,
             f"{ST_RECON} total", present=has_s, absent=blank_s)
        line("Total Staff Outstanding",
             f"={self.recon_total_cells.get('outstanding', '')}" if has_s else None,
             f"{ST_RECON} total", present=has_s, absent=blank_s)
        staff_out = f"B{row - 1}" if has_s else None

        combined_cost = " + ".join(x for x in (karigar_earned, staff_earned) if x)
        row = S.write_row(ws, self.style, row,
                          ["Combined Labour Cost (Karigar + Staff)",
                           f"={combined_cost}" if combined_cost else None,
                           "the two Earning lines above"],
                          [None, S.MONEY, None], subtotal=True)
        combined_cell = f"B{row - 1}"
        combined_out = " + ".join(x for x in (karigar_out, staff_out) if x)
        row = S.write_row(ws, self.style, row,
                          ["Combined Outstanding (Karigar + Staff)",
                           f"={combined_out}" if combined_out else None,
                           "the two Outstanding lines above"],
                          [None, S.MONEY, None], subtotal=True)

        line("Total Pieces Produced (Karigar)",
             f"={k.get('pieces', '')}" if has_k else None,
             f"{K_EARNINGS} grand total", fmt=S.QTY, present=has_k, absent=blank_k)
        pieces_cell = f"B{row - 1}" if has_k else None

        if has_k:
            row = S.write_row(
                ws, self.style, row,
                ["Overall Cost per Piece — stitching only",
                 f"=IF({pieces_cell}=0,\"\",{karigar_earned}/{pieces_cell})",
                 "Karigar Production Value / Pieces"],
                [None, S.MONEY, None])

        row += 1
        row = S.note(ws, self.style, row,
                     "§4 — the two cost bases are shown side by side rather than "
                     "blended. Staff overhead (cutting, QC, ironing, packing) and "
                     "karigar stitching answer different questions, and a single "
                     "per-piece number that mixes them answers neither. Add the "
                     "blended row only on the owner's instruction.")
        if not has_k:
            row = S.note(ws, self.style, row, f"{blank_k}.")
        if not has_s:
            row = S.note(ws, self.style, row, f"{blank_s}.")

        # What recalc.py must find here.
        if has_k:
            self.out.expect[f"{OVERVIEW}!{karigar_earned}"] = \
                self.inp.karigar.totals["earned"]
            self.out.expect[f"{OVERVIEW}!{pieces_cell}"] = \
                self.inp.karigar.totals["pieces"]
        if has_s:
            self.out.expect[f"{OVERVIEW}!{staff_earned}"] = \
                round(self.inp.payroll["total"], 2)
        if has_k and has_s:
            self.out.expect[f"{OVERVIEW}!{combined_cell}"] = round(
                self.inp.karigar.totals["earned"] + self.inp.payroll["total"], 2)

    # =======================================================================
    # KARIGAR SHEETS  (§2.4)
    # =======================================================================

    def karigar_sheets(self):
        k = self.inp.karigar
        if k is None:
            for name in (K_ITEMWISE, K_COMBINED, K_EARNINGS, K_DETAIL,
                         K_PAYSUMMARY, K_PAYDETAIL, K_RAW):
                ws = self.sheet(name)
                S.prepare(ws, self.style)
                ws.column_dimensions["A"].width = 90
                r = S.title(ws, self.style, 1, name)
                S.note(ws, self.style, r, "No karigar data uploaded for this FY.")
                self.out.skipped.append(name)
            return
        self.k_raw()
        self.k_detail()
        self.k_itemwise()
        self.k_combined()
        self.k_earnings()
        self.k_payment_summary()
        self.k_payment_detail()

    def k_raw(self):
        """§2.4.8 — every parsed row, with a Period column. The audit trail that
        every other karigar sheet's formulas point back at."""
        ws = self.sheet(K_RAW)
        S.prepare(ws, self.style)
        row = S.title(ws, self.style, 1, f"Raw Transactions (Audit) — FY{self.fy}",
                      "One row per non-null production cell in the source, exactly as "
                      "parsed. Every other karigar sheet is computed from this one.")
        row = S.header(ws, self.style, row,
                       ["Period", "Karigar", "Design", "Component", "Qty", "Rate",
                        "Value", "Source"],
                       [12, 26, 26, 22, 10, 10, 14, 34])
        start = row
        for e in self.inp.karigar.entries:
            row = S.write_row(ws, self.style, row, [
                str(e.extra.get("period") or ""),
                self._label(e.who),
                e.what or "",
                e.set_type or "",
                e.qty,
                e.rate,
                f"=E{row}*F{row}" if e.rate else e.value,
                e.where,
            ], [None, None, None, None, S.QTY, S.MONEY, S.MONEY, None])
        self.k_raw_rows = (start, max(row - 1, start))
        total = S.write_row(ws, self.style, row, [
            "TOTAL", "", "", "",
            S.total("E", start, row - 1), "",
            S.total("G", start, row - 1), "",
        ], [None, None, None, None, S.QTY, None, S.MONEY, None], subtotal=True)
        self.k_totals["raw_value"] = f"{S.quote(K_RAW)}!G{row}"
        self.k_totals["raw_qty"] = f"{S.quote(K_RAW)}!E{row}"
        self.out.expect[f"{K_RAW}!G{row}"] = self.inp.karigar.totals["earned"]
        self.out.expect[f"{K_RAW}!E{row}"] = self.inp.karigar.totals["pieces"]

    def _label(self, unit) -> str:
        return str(unit or "")

    def k_detail(self):
        """§2.4.5 — component lines, grouped by karigar."""
        ws = self.sheet(K_DETAIL)
        S.prepare(ws, self.style)
        row = S.title(ws, self.style, 1, f"Karigar x Design Detail — FY{self.fy}",
                      "Component lines grouped by the paying unit. Value is Qty x Rate "
                      "on every line; the karigar subtotal is the sum of its own lines.")
        row = S.header(ws, self.style, row,
                       ["Karigar", "Design", "Component", "Period", "Qty", "Rate",
                        "Value"],
                       [26, 26, 22, 12, 10, 10, 14])
        by_unit = defaultdict(list)
        for e in self.inp.karigar.entries:
            by_unit[self._label(e.who)].append(e)

        self.k_detail_unit_rows = {}
        start_all = row
        for unit in sorted(by_unit):
            first = row
            for e in by_unit[unit]:
                row = S.write_row(ws, self.style, row, [
                    unit, e.what or "", e.set_type or "",
                    str(e.extra.get("period") or ""),
                    e.qty, e.rate,
                    f"=E{row}*F{row}" if e.rate else e.value,
                ], [None, None, None, None, S.QTY, S.MONEY, S.MONEY])
            row = S.write_row(ws, self.style, row, [
                f"{unit} — subtotal", "", "", "",
                S.total("E", first, row - 1), "",
                S.total("G", first, row - 1),
            ], [None, None, None, None, S.QTY, None, S.MONEY], subtotal=True)
            self.k_detail_unit_rows[unit] = (first, row - 2)
            row += 1
        self.k_detail_rows = (start_all, row - 1)

    def k_itemwise(self):
        """§2.4.2 — component lines per design, Qty x Rate = Value."""
        ws = self.sheet(K_ITEMWISE)
        S.prepare(ws, self.style)
        row = S.title(ws, self.style, 1, f"Item-wise Production & Cost — FY{self.fy}",
                      "One line per design and component. Value is a formula on every "
                      "line, and each design's lines sum to its recorded total.")
        row = S.header(ws, self.style, row,
                       ["Design", "Component", "Qty", "Weighted Rate", "Value"],
                       [28, 24, 12, 14, 16])
        pooled = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))
        for e in self.inp.karigar.entries:
            slot = e.set_type or ""
            bucket = pooled[e.what or ""][slot]
            bucket[0] += e.qty
            bucket[1] += float(e.value or 0.0)
        start = row
        for design in sorted(pooled):
            for component in sorted(pooled[design]):
                qty, value = pooled[design][component]
                rate = (value / qty) if qty else 0.0
                row = S.write_row(ws, self.style, row,
                                  [design, component, qty, rate, f"=C{row}*D{row}"],
                                  [None, None, S.QTY, S.RATE, S.MONEY])
        total_row = row
        row = S.write_row(ws, self.style, row,
                          ["TOTAL", "", S.total("C", start, row - 1), "",
                           S.total("E", start, row - 1)],
                          [None, None, S.QTY, None, S.MONEY], subtotal=True)
        self.out.expect[f"{K_ITEMWISE}!E{total_row}"] = self.inp.karigar.totals["earned"]

    def k_combined(self):
        """§2.4.3 — quantities only, and the set-matching result per design."""
        ws = self.sheet(K_COMBINED)
        S.prepare(ws, self.style)
        row = S.title(ws, self.style, 1, f"Combined Production — FY{self.fy}",
                      "Quantities only, no money. Complete Sets is the smallest "
                      "POPULATED slot (§2.2); the surplus columns show what is left "
                      "over, unmerged.")
        row = S.header(ws, self.style, row,
                       ["Design", "Set Type calls for", "Top", "Bottom", "Dupatta",
                        "Complete Sets", "Extra Top", "Extra Bottom",
                        "Pending Dupatta", "Extra Dupatta"],
                       [28, 24, 10, 10, 10, 14, 12, 13, 15, 14])
        start = row
        designs = self.inp.karigar.designs
        for name in sorted(designs):
            d = designs[name]
            slots = d.slots
            surplus = d.surplus
            required = ", ".join(d.required) if d.required else "(as produced)"
            extra_top = surplus.get(TOP, 0)
            extra_bottom = surplus.get(BOTTOM, 0)
            pending = d.pending_dupatta
            row = S.write_row(ws, self.style, row, [
                name, required,
                slots.get(TOP, 0), slots.get(BOTTOM, 0), slots.get(DUPATTA, 0),
                d.complete_sets,
                extra_top - pending, extra_bottom - pending, pending,
                d.extra_dupatta,
            ], [None, None, S.INT, S.INT, S.INT, S.INT, S.INT, S.INT, S.INT, S.INT])
        total_row = row
        row = S.write_row(ws, self.style, row, [
            "TOTAL", "",
            S.total("C", start, row - 1), S.total("D", start, row - 1),
            S.total("E", start, row - 1), S.total("F", start, row - 1),
            S.total("G", start, row - 1), S.total("H", start, row - 1),
            S.total("I", start, row - 1), S.total("J", start, row - 1),
        ], [None, None] + [S.INT] * 8, subtotal=True)
        self.out.expect[f"{K_COMBINED}!F{total_row}"] = \
            self.inp.karigar.totals["complete_sets"]

    def k_earnings(self):
        """§2.4.4 — formula-linked to the Karigar x Design Detail sheet."""
        ws = self.sheet(K_EARNINGS)
        S.prepare(ws, self.style)
        row = S.title(ws, self.style, 1, f"Karigar Earnings — FY{self.fy}",
                      f"Every figure is a SUMIFS against '{K_DETAIL}'. Nothing here is "
                      "typed in, so a corrected detail line moves the earnings the "
                      "moment it is corrected.")
        row = S.header(ws, self.style, row,
                       ["Karigar", "Pieces", "Earned", "Paid", "Outstanding"],
                       [30, 14, 16, 16, 16])
        start = row
        d0, d1 = self.k_detail_rows
        units = self.inp.karigar.units
        for unit in sorted(units):
            label = self._label(unit)
            arg = S.text_arg(label)
            row = S.write_row(ws, self.style, row, [
                label,
                f"=SUMIFS({S.quote(K_DETAIL)}!$E${d0}:$E${d1},"
                f"{S.quote(K_DETAIL)}!$A${d0}:$A${d1},{arg})",
                f"=SUMIFS({S.quote(K_DETAIL)}!$G${d0}:$G${d1},"
                f"{S.quote(K_DETAIL)}!$A${d0}:$A${d1},{arg})",
                f"=SUMIFS({S.quote(K_PAYSUMMARY)}!$C$1:$C$5000,"
                f"{S.quote(K_PAYSUMMARY)}!$A$1:$A$5000,{arg})",
                f"=C{row}-D{row}",
            ], [None, S.QTY, S.MONEY, S.MONEY, S.MONEY])
        gt = row
        row = S.write_row(ws, self.style, row, [
            "GRAND TOTAL",
            S.total("B", start, row - 1), S.total("C", start, row - 1),
            S.total("D", start, row - 1), S.total("E", start, row - 1),
        ], [None, S.QTY, S.MONEY, S.MONEY, S.MONEY], subtotal=True)
        self.k_totals["pieces"] = f"{S.quote(K_EARNINGS)}!B{gt}"
        self.k_totals["earned"] = f"{S.quote(K_EARNINGS)}!C{gt}"
        self.k_totals["paid"] = f"{S.quote(K_EARNINGS)}!D{gt}"
        self.k_totals["outstanding"] = f"{S.quote(K_EARNINGS)}!E{gt}"
        t = self.inp.karigar.totals
        self.out.expect[f"{K_EARNINGS}!B{gt}"] = t["pieces"]
        self.out.expect[f"{K_EARNINGS}!C{gt}"] = t["earned"]
        self.out.expect[f"{K_EARNINGS}!D{gt}"] = t["paid"]
        self.out.expect[f"{K_EARNINGS}!E{gt}"] = t["outstanding"]

    def k_payment_summary(self):
        """§2.4.6 — per unit, per period, and the Combined columns as formulas.

        §7 checks that the Combined columns equal the sum of each period's. They
        are written as that sum, so the check cannot be passed by accident and
        cannot be failed by a typo.
        """
        ws = self.sheet(K_PAYSUMMARY)
        S.prepare(ws, self.style)
        row = S.title(ws, self.style, 1, f"Payment Summary (Yearly) — FY{self.fy}",
                      "The Combined columns are the sum of the period columns beside "
                      "them, written as formulas — never read from the source.")
        k = self.inp.karigar
        periods = sorted({p for u in k.by_period.values() for p in u} |
                         {p for u in k.paid_by_period.values() for p in u})
        head = ["Karigar", "Combined Earned", "Combined Paid", "Combined Outstanding"]
        for p in periods:
            head += [f"{p} Earned", f"{p} Paid", f"{p} Outstanding"]
        row = S.header(ws, self.style, row, head,
                       [30, 16, 16, 18] + [15] * (3 * len(periods)))
        start = row
        units = sorted(set(k.by_period) | set(k.paid_by_period))
        for unit in units:
            label = self._label(unit)
            earned_cells, paid_cells, out_cells = [], [], []
            values = [label, None, None, None]
            for i, p in enumerate(periods):
                base = 5 + i * 3
                earned_cells.append(f"{S.col(base)}{row}")
                paid_cells.append(f"{S.col(base + 1)}{row}")
                out_cells.append(f"{S.col(base + 2)}{row}")
                values += [
                    round(k.by_period.get(unit, {}).get(p, 0.0), 2),
                    round(k.paid_by_period.get(unit, {}).get(p, 0.0), 2),
                    f"={S.col(base)}{row}-{S.col(base + 1)}{row}",
                ]
            values[1] = "=" + "+".join(earned_cells) if earned_cells else 0
            values[2] = "=" + "+".join(paid_cells) if paid_cells else 0
            values[3] = f"=B{row}-C{row}"
            row = S.write_row(ws, self.style, row, values,
                              [None] + [S.MONEY] * (3 + 3 * len(periods)))
        gt = row
        row = S.write_row(
            ws, self.style, row,
            ["GRAND TOTAL"] + [S.total(S.col(c), start, row - 1)
                               for c in range(2, 5 + 3 * len(periods))],
            [None] + [S.MONEY] * (3 + 3 * len(periods)), subtotal=True)
        self.out.expect[f"{K_PAYSUMMARY}!B{gt}"] = k.totals["earned"]
        self.out.expect[f"{K_PAYSUMMARY}!C{gt}"] = k.totals["paid"]
        self.out.expect[f"{K_PAYSUMMARY}!D{gt}"] = k.totals["outstanding"]

    def k_payment_detail(self):
        """§2.4.7 — monthly, wherever payment dates exist."""
        ws = self.sheet(K_PAYDETAIL)
        S.prepare(ws, self.style)
        row = S.title(ws, self.style, 1, f"Payment Detail (Monthly) — FY{self.fy}")
        dated = [e for e in self.inp.karigar.entries
                 if e.date is not None and e.paid is not None]
        if not dated:
            S.note(ws, self.style, row,
                   "The payment data for this FY carries no dates — only a name and "
                   "an amount — so a month-wise split is not possible. Payment "
                   f"Summary (Yearly) reconciles at FY level instead. When the source "
                   "adopts Date | Name | Amount | Method | Note, this sheet fills "
                   "itself in with no change to the rules.")
            self.out.notes.append("karigar payment detail: no dated payments in source")
            return
        row = S.header(ws, self.style, row,
                       ["Date", "Month", "Karigar", "Amount", "Source"],
                       [14, 12, 30, 16, 34])
        start = row
        for e in sorted(dated, key=lambda x: str(x.date)):
            row = S.write_row(ws, self.style, row, [
                e.date, f'=TEXT(A{row},"mmm-yyyy")', self._label(e.who),
                e.paid, e.where,
            ], [S.DATE, None, None, S.MONEY, None])
        S.write_row(ws, self.style, row,
                    ["TOTAL", "", "", S.total("D", start, row - 1), ""],
                    [None, None, None, S.MONEY, None], subtotal=True)

    # =======================================================================
    # STAFF SHEETS  (§3.6)
    # =======================================================================

    def staff_sheets(self):
        if not self.inp.has_staff:
            for name in (ST_SALARY, ST_THRESHOLD, ST_MASTER, ST_ATTENDANCE,
                         ST_MONTHLY, ST_PAYDETAIL, ST_RECON, ST_WORK,
                         ST_PERFORMANCE):
                ws = self.sheet(name)
                S.prepare(ws, self.style)
                ws.column_dimensions["A"].width = 90
                r = S.title(ws, self.style, 1, name)
                S.note(ws, self.style, r, "No staff data uploaded for this FY.")
                self.out.skipped.append(name)
            return
        self.s_salary_log()
        self.s_threshold_log()
        self.s_master()
        self.s_attendance()
        self.s_work_report()
        self.s_monthly()
        self.s_payment_detail()
        self.s_reconciliation()
        self.s_performance()

    def _log_sheet(self, name, title_text, note_text, log, value_header, fmt):
        ws = self.sheet(name)
        S.prepare(ws, self.style)
        row = S.title(ws, self.style, 1, title_text, note_text)
        row = S.header(ws, self.style, row,
                       ["Staff", "Effective From", "Effective To (blank = ongoing)",
                        "Effective To (calc)", value_header],
                       [22, 16, 26, 18, 18])
        start = row
        for key in sorted(log.keys()):
            for r in log.rows(key):
                if r.value is None:
                    continue
                row = S.write_row(ws, self.style, row, [
                    key,
                    r.frm,
                    r.to,
                    f"=IF(C{row}=\"\",{S.FOREVER},C{row})",
                    float(r.value),
                ], [None, S.DATE, S.DATE, S.DATE, fmt])
        return ws, (start, max(row - 1, start))

    def s_salary_log(self):
        _, rows = self._log_sheet(
            ST_SALARY, f"Salary Log — FY{self.fy}",
            "Effective-dated. A raise appends a row and closes the one before it; "
            "history is never rewritten, so any past month still resolves to what "
            "was actually in force then.",
            self.inp.master.salary, "Monthly Salary (Rs)", S.MONEY)
        self.salary_rows = rows

    def s_threshold_log(self):
        _, rows = self._log_sheet(
            ST_THRESHOLD, f"Threshold Log — FY{self.fy}",
            "Tracked independently of salary, with the same effective-dating "
            "mechanics. A threshold change and a raise on different dates are two "
            "logs, not one.",
            self.inp.master.threshold_days, "Threshold Days/Mo", S.INT)
        self.threshold_rows = rows

    def s_master(self):
        """§3.6.3 — plus the hours reference table §3.6.4 does its INDEX/MATCH on."""
        ws = self.sheet(ST_MASTER)
        S.prepare(ws, self.style)
        row = S.title(ws, self.style, 1, f"Staff Master — FY{self.fy}")

        row = S.write_row(ws, self.style, row,
                          ["Hours Reference (§3.5) — informational only, never prices "
                           "a month"], bold=True)
        hdr = S.header(ws, self.style, row, ["Gender", "Weekday", "Sunday"],
                       [22, 14, 14], band=2)
        self.hours_ref_rows = (hdr, hdr + 1)
        sh = self.inp.master.shift_hours
        S.write_row(ws, self.style, hdr,
                    ["M", sh.get(("M", "Weekday"), 10.0), sh.get(("M", "Sunday"), 5.0)],
                    [None, S.QTY, S.QTY])
        row = S.write_row(ws, self.style, hdr + 1,
                          ["F", sh.get(("F", "Weekday"), 8.0), sh.get(("F", "Sunday"), 5.5)],
                          [None, S.QTY, S.QTY])
        row = S.note(ws, self.style, row + 1,
                     "A half-day is half of that gender and day's present hours. "
                     "Absent, holiday and blank are zero.")
        row += 1

        row = S.header(ws, self.style, row,
                       ["Name", "Gender", "Threshold Hrs/Mo (legacy reference only)",
                        "Roles", "Status", "Pay Basis", "Blended FY Daily Rate",
                        "Blended FY Hourly Rate"],
                       [22, 10, 26, 30, 12, 16, 20, 20])
        ws.freeze_panes = ws.cell(row=row, column=1)
        start = row
        m = self.inp.master
        for staff in self.staff_order():
            p = m.person(staff)
            basis = self.any_basis(staff)
            legacy = m.threshold_hours.maybe(staff, self.months[0])
            if legacy is None:
                for mm in self.months:
                    legacy = m.threshold_hours.maybe(staff, mm)
                    if legacy is not None:
                        break
            if PIECE_RATE in basis:
                # No daily rate exists to average. §3.5 gives these staff a flat
                # rate outright, and pointing the Work Report at a blank average
                # would cost their hours at nothing.
                daily = None
                hourly = round(blended_hourly(m, staff, self.fy), 4)
            else:
                daily = (f"=IFERROR(AVERAGEIFS({S.quote(ST_MONTHLY)}!$L$1:$L$5000,"
                         f"{S.quote(ST_MONTHLY)}!$A$1:$A$5000,{S.text_arg(staff)},"
                         f"{S.quote(ST_MONTHLY)}!$L$1:$L$5000,\">0\"),\"\")")
                hourly = (f"=IFERROR(G{row}/INDEX($B${self.hours_ref_rows[0]}:"
                          f"$B${self.hours_ref_rows[1]},"
                          f"MATCH(B{row},$A${self.hours_ref_rows[0]}:"
                          f"$A${self.hours_ref_rows[1]},0)),\"\")")
            row = S.write_row(ws, self.style, row, [
                staff, p.group, legacy, ", ".join(p.roles), p.roster, basis,
                daily, hourly,
            ], [None, None, S.INT, None, None, None, S.RATE, S.RATE])
        self.master_rows = (start, max(row - 1, start))

        row = S.note(ws, self.style, row + 1,
                     "Threshold Hrs/Mo is carried for reference only (§3.6.3). Nothing "
                     "prices off it: the daily rate is the salary over the threshold "
                     "DAYS, and the hourly rate above is that daily rate over this "
                     "person's own weekday shift.")
        S.note(ws, self.style, row,
               "The blended daily rate averages the months this person was actually "
               "employed. A month they were not employed has no daily rate to average.")

        for staff in self.staff_order():
            want = blended_hourly(m, staff, self.fy)
            if want:
                idx = self.staff_order().index(staff)
                self.out.expect[f"{ST_MASTER}!H{start + idx}"] = round(want, 4)

    def s_attendance(self):
        """§3.6.4 — the raw marks, and the reference hours beside each of them."""
        ws = self.sheet(ST_ATTENDANCE)
        S.prepare(ws, self.style)
        row = S.title(ws, self.style, 1, f"Attendance — FY{self.fy}",
                      "Status as recorded. Hours are a reference column, computed from "
                      "the gender on Staff Master and the day of the week — they never "
                      "price a month.")
        # §3.5 — piece-rate staff never appear on an attendance sheet.
        tracked = [s for s in self.staff_order() if PIECE_RATE not in self.any_basis(s)]

        name_row = row
        head_row = row + 1
        ws.cell(row=name_row, column=1, value="").font = self.style.body_font
        labels = ["Date", "Day"]
        for i, staff in enumerate(tracked):
            c = ws.cell(row=name_row, column=3 + i * 2, value=staff)
            c.font = self.style.header_font
            c.fill = self.style.header_fill_2
            c.alignment = self.style.centre
            c2 = ws.cell(row=name_row, column=4 + i * 2, value="")
            c2.fill = self.style.header_fill_2
            ws.merge_cells(start_row=name_row, start_column=3 + i * 2,
                           end_row=name_row, end_column=4 + i * 2)
            labels += ["Status", "Hours"]
            self.att_cols[staff] = (S.col(3 + i * 2), S.col(4 + i * 2))
        row = S.header(ws, self.style, head_row, labels,
                       [14, 12] + [9, 9] * len(tracked))
        ws.freeze_panes = ws.cell(row=head_row + 1, column=3)

        hr0, hr1 = self.hours_ref_rows
        mr0, mr1 = self.master_rows
        start = row
        book = self.inp.book
        days = [d for mth in self.months for d in mth.dates()]
        for d in days:
            values = [d, f'=TEXT(A{row},"ddd")']
            for staff in tracked:
                code = book.get(staff, d) if book is not None else None
                status = S.col(3 + tracked.index(staff) * 2)
                values.append(code or "")
                values.append(
                    f'=IF(OR({status}{row}="",{status}{row}="A",{status}{row}="HL"),0,'
                    f'IF({status}{row}="H",0.5,1)*'
                    f'INDEX({S.quote(ST_MASTER)}!$B${hr0}:$C${hr1},'
                    f'MATCH(INDEX({S.quote(ST_MASTER)}!$B${mr0}:$B${mr1},'
                    f'MATCH({status}${name_row},{S.quote(ST_MASTER)}!$A${mr0}:$A${mr1},0)),'
                    f'{S.quote(ST_MASTER)}!$A${hr0}:$A${hr1},0),'
                    f'IF(WEEKDAY($A{row},1)=1,2,1)))'
                )
            row = S.write_row(ws, self.style, row, values,
                              [S.DATE, None] + [None, S.QTY] * len(tracked))
        self.att_rows = (start, row - 1)

    def s_work_report(self):
        """§3.6.8 — raw design/role hours as imported, costed at the blended rate.

        The rate-reference row sits directly under the header and pulls each
        column's rate from Staff Master, so Total Labour Cost is one SUMPRODUCT
        and every rate on the sheet is visible rather than buried in a formula.
        """
        ws = self.sheet(ST_WORK)
        S.prepare(ws, self.style)
        row = S.title(
            ws, self.style, 1, f"Work Report — Cost per Design, FY{self.fy}",
            "Labour cost only. Stitching is the karigar pipeline and appears nowhere "
            "on this sheet — adding the two would count the same garment twice.")

        pairs = sorted({(r.staff, r.role) for r in self.inp.work_rows})
        designs = sorted({r.design for r in self.inp.work_rows})
        hours = defaultdict(float)
        for r in self.inp.work_rows:
            hours[(r.design, r.staff, r.role)] += r.hours

        head = ["Design", "Quantity"] + [f"{s} — {role}" if role else s
                                         for s, role in pairs] + \
               ["Total Hours", "Total Labour Cost", "Cost per Piece"]
        head_row = row
        row = S.header(ws, self.style, row, head,
                       [28, 12] + [16] * len(pairs) + [14, 18, 16])

        first_pair = 3
        last_pair = 2 + len(pairs)
        rate_row = row
        mr0, mr1 = self.master_rows
        rate_values = ["Rs/hr →", None]
        for staff, _role in pairs:
            rate_values.append(
                f"=IFERROR(ROUND(INDEX({S.quote(ST_MASTER)}!$H${mr0}:$H${mr1},"
                f"MATCH({S.text_arg(staff)},{S.quote(ST_MASTER)}!$A${mr0}:$A${mr1},0)),4),0)")
        rate_values += [None, None, None]
        row = S.write_row(ws, self.style, row, rate_values,
                          [None, None] + [S.RATE] * len(pairs), bold=True)
        for c in range(1, len(head) + 1):
            ws.cell(row=rate_row, column=c).fill = self.style.accent_fill
        ws.freeze_panes = ws.cell(row=rate_row + 1, column=1)

        pr_first, pr_last = S.col(first_pair), S.col(last_pair)
        th_col = S.col(last_pair + 1)
        tc_col = S.col(last_pair + 2)
        cpp_col = S.col(last_pair + 3)

        start = row
        for design in designs:
            values = [design, float(self.inp.quantities.get(design, 0) or 0)]
            for staff, role in pairs:
                got = hours.get((design, staff, role), 0.0)
                values.append(round(got, 4) if got else None)
            values += [
                f"=SUM({pr_first}{row}:{pr_last}{row})",
                f"=ROUND(SUMPRODUCT({pr_first}{row}:{pr_last}{row},"
                f"{pr_first}${rate_row}:{pr_last}${rate_row}),2)",
                f'=IF(B{row}=0,"",{tc_col}{row}/B{row})',
            ]
            row = S.write_row(ws, self.style, row, values,
                              [None, S.QTY] + [S.QTY] * len(pairs)
                              + [S.QTY, S.MONEY, S.MONEY])
        data_end = row - 1
        row = S.write_row(
            ws, self.style, row,
            ["TOTAL", S.total("B", start, data_end)]
            + [S.total(S.col(c), start, data_end)
               for c in range(first_pair, last_pair + 1)]
            + [S.total(th_col, start, data_end),
               S.total(tc_col, start, data_end), ""],
            [None, S.QTY] + [S.QTY] * len(pairs) + [S.QTY, S.MONEY, None],
            subtotal=True)
        self.work_total_cost_cell = f"{S.quote(ST_WORK)}!{tc_col}{row - 1}"
        if self.inp.allocation:
            self.out.expect[f"{ST_WORK}!{tc_col}{row - 1}"] = \
                self.inp.allocation["allocated_cost"]
            self.out.expect[f"{ST_WORK}!{th_col}{row - 1}"] = \
                round(self.inp.allocation["logged_hours"], 4)

        row += 1
        row = S.write_row(ws, self.style, row,
                          ["Per staff, this FY — hours logged and what they cost"],
                          bold=True)
        block = S.header(ws, self.style, row, ["Staff", "Rs/hr", "FY Hours", "FY Cost"],
                         None, band=2)
        ws.freeze_panes = ws.cell(row=rate_row + 1, column=1)
        by_staff = defaultdict(list)
        for i, (staff, _role) in enumerate(pairs):
            by_staff[staff].append(S.col(first_pair + i))
        r = block
        for staff in sorted(by_staff):
            cols = by_staff[staff]
            hours_sum = ("+".join(f"SUM({c}${start}:{c}${data_end})" for c in cols)
                         if data_end >= start else "0")
            rate_ref = f"{cols[0]}${rate_row}"
            S.write_row(ws, self.style, r, [
                staff, f"={rate_ref}", f"={hours_sum}", f"=B{r}*C{r}",
            ], [None, S.RATE, S.QTY, S.MONEY])
            self.work_staff_rows[staff] = r
            r += 1
        S.note(ws, self.style, r + 1,
               "A piece-rate staff member's whole wage for the year is their FY Cost "
               "line here (§3.5). The Work Report is the only place their hours are "
               "recorded, so it is the only place their wage can come from.")

    def s_monthly(self):
        """§3.6.5 — the sheet everything else reconciles to.

        Every count is a COUNTIFS against the raw Attendance grid, every rate a
        SUMIFS against the effective-dated logs, and the earning branches on the
        pay basis exactly as §3.5 sets out. Nothing is carried across from the
        engine except the marks themselves.
        """
        ws = self.sheet(ST_MONTHLY)
        S.prepare(ws, self.style)
        row = S.title(ws, self.style, 1, f"Monthly Summary — FY{self.fy}",
                      "Twelve rows per person and an FY subtotal. Every figure is a "
                      "formula over the Attendance grid and the two logs.")
        row = S.header(ws, self.style, row,
                       ["Staff", "Month", "From", "To", "P", "H", "A (incl. blank)",
                        "HL", "Actual Days-Equivalent", "Threshold Days",
                        "Resolved Monthly Salary", "Daily Rate", "Earning",
                        "Hours Worked", "State"],
                       [20, 12, 12, 12, 7, 7, 14, 7, 20, 15, 20, 14, 16, 14, 14])
        start = row
        a0, a1 = self.att_rows
        s0, s1 = self.salary_rows
        t0, t1 = self.threshold_rows
        mr0, mr1 = self.master_rows
        att = S.quote(ST_ATTENDANCE)
        sal = S.quote(ST_SALARY)
        thr = S.quote(ST_THRESHOLD)
        mas = S.quote(ST_MASTER)

        for staff in self.staff_order():
            arg = S.text_arg(staff)
            basis = self.any_basis(staff)
            first = row
            if PIECE_RATE in basis:
                # §3.2.2 — the Work Report is a whole-FY aggregate with no date
                # column, so a piece-rate wage genuinely has no month inside it.
                # Twelve invented rows would be twelve lies.
                wr = self.work_staff_rows.get(staff)
                earning = (f"={S.quote(ST_WORK)}!$D${wr}" if wr else 0)
                row = S.write_row(ws, self.style, row, [
                    staff, NO_MONTH_BREAKDOWN, None, None, None, None, None, None,
                    None, None, None, None, earning, None, "Piece-rate",
                ], [None] * 12 + [S.MONEY, None, None])
                row = S.write_row(ws, self.style, row, [
                    staff, FY_TOTAL, None, None, None, None, None, None, None, None,
                    None, None, f"=M{row - 1}", None, "",
                ], [None] * 12 + [S.MONEY, None, None], subtotal=True)
                continue

            status_col = self.att_cols.get(staff, (None, None))[0]
            hours_col = self.att_cols.get(staff, (None, None))[1]
            for mth in self.months:
                frm, to = mth.first_day, mth.last_day
                state = self._state_of(staff, mth)
                if status_col:
                    counts = [
                        self._countifs(att, status_col, a0, a1, row, code)
                        for code in ("P", "H", "A", "HL")
                    ]
                    # A blank cell inside an employment spell is an absence for
                    # pay (§3.2) even though it is a tracking gap for performance.
                    counts[2] = (
                        f'={self._countifs_body(att, status_col, a0, a1, row, "A")}'
                        f'+{self._countifs_body(att, status_col, a0, a1, row, "")}'
                    ) if state != NOT_EMPLOYED else None
                    hours_worked = (
                        f"=SUMIFS({att}!${hours_col}${a0}:${hours_col}${a1},"
                        f"{att}!$A${a0}:$A${a1},\">=\"&$C{row},"
                        f"{att}!$A${a0}:$A${a1},\"<=\"&$D{row})"
                    )
                else:
                    counts, hours_worked = [None] * 4, None

                threshold = (f"=SUMIFS({thr}!$E${t0}:$E${t1},{thr}!$A${t0}:$A${t1},$A{row},"
                             f"{thr}!$B${t0}:$B${t1},\"<=\"&$C{row},"
                             f"{thr}!$D${t0}:$D${t1},\">=\"&$C{row})")
                salary = (f"=SUMIFS({sal}!$E${s0}:$E${s1},{sal}!$A${s0}:$A${s1},$A{row},"
                          f"{sal}!$B${s0}:$B${s1},\"<=\"&$C{row},"
                          f"{sal}!$D${s0}:$D${s1},\">=\"&$C{row})")
                daily = f'=IF(J{row}=0,0,K{row}/J{row})'
                if state == NOT_EMPLOYED:
                    counts = [None] * 4
                    threshold = salary = daily = None
                    days_equiv = None
                    earning = 0
                    hours_worked = None
                else:
                    days_equiv = f"=E{row}+H{row}+0.5*F{row}"
                    earning = (
                        f'=IF(INDEX({mas}!$F${mr0}:$F${mr1},'
                        f'MATCH($A{row},{mas}!$A${mr0}:$A${mr1},0))="{FLAT}",'
                        f'K{row},L{row}*I{row})'
                    )
                row = S.write_row(ws, self.style, row, [
                    staff, mth.key, frm, to,
                    counts[0], counts[1], counts[2], counts[3],
                    days_equiv, threshold, salary, daily, earning, hours_worked,
                    state,
                ], [None, None, S.DATE, S.DATE, S.INT, S.INT, S.INT, S.INT,
                    S.QTY, S.INT, S.MONEY, S.RATE, S.MONEY, S.QTY, None])
            row = S.write_row(ws, self.style, row, [
                staff, FY_TOTAL, None, None,
                S.total("E", first, row - 1), S.total("F", first, row - 1),
                S.total("G", first, row - 1), S.total("H", first, row - 1),
                S.total("I", first, row - 1), None, None, None,
                S.total("M", first, row - 1), S.total("N", first, row - 1), "",
            ], [None, None, None, None, S.INT, S.INT, S.INT, S.INT, S.QTY, None,
                None, None, S.MONEY, S.QTY, None], subtotal=True)
            want = round(sum(r.earning for r in self.rows_of(staff)), 2)
            self.out.expect[f"{ST_MONTHLY}!M{row - 1}"] = want
        self.monthly_rows = (start, row - 1)

    def _state_of(self, staff, month) -> str:
        for r in self.inp.payroll.get("rows", []):
            if r.staff == staff and r.month.key == month.key:
                return r.state
        return NOT_EMPLOYED

    def _countifs_body(self, att, col, a0, a1, row, code) -> str:
        return (f"COUNTIFS({att}!$A${a0}:$A${a1},\">=\"&$C{row},"
                f"{att}!$A${a0}:$A${a1},\"<=\"&$D{row},"
                f"{att}!${col}${a0}:${col}${a1},{S.text_arg(code)})")

    def _countifs(self, att, col, a0, a1, row, code) -> str:
        return "=" + self._countifs_body(att, col, a0, a1, row, code)

    def s_payment_detail(self):
        """§3.6.6 — all five columns now, so the sheet is ready the day dates arrive."""
        ws = self.sheet(ST_PAYDETAIL)
        S.prepare(ws, self.style)
        row = S.title(ws, self.style, 1, f"Payment Detail — FY{self.fy}",
                      "Date, Method and Note are carried now even though the source "
                      "supplies only Name and Amount. When the source adopts the fuller "
                      "format the sheet fills in with no change to the rules.")
        row = S.header(ws, self.style, row,
                       ["Date", "Name", "Amount", "Method", "Note"],
                       [14, 22, 16, 16, 40])
        start = row
        for staff in sorted(self.inp.payments):
            row = S.write_row(ws, self.style, row, [
                None, staff, round(self.inp.payments[staff], 2), None,
                "no date in source — reconciled at FY level (§3.2.3)",
            ], [S.DATE, None, S.MONEY, None, None])
        if row == start:
            S.note(ws, self.style, row, "No payment data was uploaded for this FY.")
            self.out.notes.append("staff payments: none supplied")
            self.pay_detail_rows = (start, start)
            return
        S.write_row(ws, self.style, row,
                    ["TOTAL", "", S.total("C", start, row - 1), "", ""],
                    [None, None, S.MONEY, None, None], subtotal=True)
        self.pay_detail_rows = (start, row - 1)

    def s_reconciliation(self):
        """§3.6.7 — FY earning against FY payments, per person."""
        ws = self.sheet(ST_RECON)
        S.prepare(ws, self.style)
        row = S.title(ws, self.style, 1, f"Payment Reconciliation — FY{self.fy}")
        note = ("Reconciled at FY total level, because the payment data carries no "
                "dates. The moment it does, this becomes month-wise with no change to "
                "the rules.") if not any(True for _ in ()) else ""
        row = S.note(ws, self.style, row, note)
        row += 1
        row = S.header(ws, self.style, row,
                       ["Staff", "FY Earning", "FY Payments", "Outstanding"],
                       [22, 18, 18, 18])
        start = row
        m0, m1 = self.monthly_rows
        p0, p1 = getattr(self, "pay_detail_rows", (1, 1))
        mon = S.quote(ST_MONTHLY)
        pay = S.quote(ST_PAYDETAIL)
        for staff in self.staff_order():
            arg = S.text_arg(staff)
            row = S.write_row(ws, self.style, row, [
                staff,
                f"=SUMIFS({mon}!$M${m0}:$M${m1},{mon}!$A${m0}:$A${m1},{arg},"
                f"{mon}!$B${m0}:$B${m1},{S.text_arg('<>' + FY_TOTAL)})",
                f"=SUMIFS({pay}!$C${p0}:$C${p1},{pay}!$B${p0}:$B${p1},{arg})",
                f"=B{row}-C{row}",
            ], [None, S.MONEY, S.MONEY, S.MONEY])
        row = S.write_row(ws, self.style, row, [
            "TOTAL", S.total("B", start, row - 1), S.total("C", start, row - 1),
            S.total("D", start, row - 1),
        ], [None, S.MONEY, S.MONEY, S.MONEY], subtotal=True)
        self.recon_total_cells = {
            "earning": f"{S.quote(ST_RECON)}!B{row - 1}",
            "paid": f"{S.quote(ST_RECON)}!C{row - 1}",
            "outstanding": f"{S.quote(ST_RECON)}!D{row - 1}",
        }
        self.out.expect[f"{ST_RECON}!B{row - 1}"] = round(self.inp.payroll["total"], 2)
        if self.inp.payments:
            self.out.expect[f"{ST_RECON}!C{row - 1}"] = \
                round(sum(self.inp.payments.values()), 2)

        S.note(ws, self.style, row + 1,
               "A Flat or Piece-rate staff member's earning is deliberately decoupled "
               "from attendance (§3.7) — their outstanding says nothing about how much "
               "they worked.")

    def s_performance(self):
        """§3.6.9 — utilisation against threshold, with the bands editable."""
        ws = self.sheet(ST_PERFORMANCE)
        S.prepare(ws, self.style)
        row = S.title(ws, self.style, 1, f"Performance Tracking — FY{self.fy}",
                      "Utilisation is Actual Days-Equivalent over Threshold Days. A "
                      "month with nothing recorded is No Data, never Below Average — a "
                      "month nobody logged is not a month somebody failed.")
        row = S.write_row(ws, self.style, row,
                          ["Rating bands — editable"], bold=True)
        bands_row = row
        ws.cell(row=row, column=1, value="Satisfactory ≥").font = self.style.body_font
        S.editable(ws, self.style, row, 2, self.inp.master.bands["satisfactory"], S.PERCENT)
        ws.cell(row=row + 1, column=1, value="Average ≥").font = self.style.body_font
        S.editable(ws, self.style, row + 1, 2, self.inp.master.bands["average"], S.PERCENT)
        ws.cell(row=row + 2, column=1,
                value="below that = Below Average").font = self.style.note_font
        row += 4

        row = S.header(ws, self.style, row,
                       ["Staff", "Month", "P", "H", "A", "HL",
                        "Actual Days-Equivalent", "Threshold Days", "Utilisation %",
                        "Rating"],
                       [20, 12, 7, 7, 7, 7, 20, 15, 14, 18])
        m0, m1 = self.monthly_rows
        mon = S.quote(ST_MONTHLY)
        start = row
        summary_rows = {}
        for staff in self.staff_order():
            arg = S.text_arg(staff)
            basis = self.any_basis(staff)
            first = row
            if PIECE_RATE in basis:
                row = S.write_row(ws, self.style, row, [
                    staff, NO_MONTH_BREAKDOWN, None, None, None, None, None, None,
                    None, "for information only — pay is not tied to attendance",
                ])
                summary_rows[staff] = (first, row - 1)
                continue
            for mth in self.months:
                src = self._monthly_row_for(staff, mth)
                state = self._state_of(staff, mth)
                if src is None or state in (NOT_EMPLOYED,):
                    row = S.write_row(ws, self.style, row, [
                        staff, mth.key, None, None, None, None, None, None, None,
                        NOT_EMPLOYED])
                    continue
                util = (f'=IF(H{row}=0,"",G{row}/H{row})')
                rating = (
                    f'=IF(OR(I{row}="",AND(C{row}=0,D{row}=0,F{row}=0,'
                    f'{mon}!$G${src}=0)),"{NO_DATA}",'
                    f'IF(I{row}>=$B${bands_row},"Satisfactory",'
                    f'IF(I{row}>=$B${bands_row + 1},"Average","Below Average")))'
                )
                row = S.write_row(ws, self.style, row, [
                    staff, mth.key,
                    f"={mon}!$E${src}", f"={mon}!$F${src}", f"={mon}!$G${src}",
                    f"={mon}!$H${src}", f"={mon}!$I${src}", f"={mon}!$J${src}",
                    util, rating,
                ], [None, None, S.INT, S.INT, S.INT, S.INT, S.QTY, S.INT,
                    S.PERCENT, None])
            summary_rows[staff] = (first, row - 1)
        perf_end = row - 1

        row += 1
        row = S.write_row(ws, self.style, row, ["Staff Performance Summary"], bold=True)
        row = S.header(ws, self.style, row,
                       ["Staff", "Avg Actual Days/Month", "Avg Utilisation",
                        "Satisfactory", "Average", "Below Average", "No Data",
                        "Overall FY Rating", "Note"],
                       [20, 20, 16, 12, 12, 14, 10, 18, 46], band=2)
        for staff in self.staff_order():
            first, last = summary_rows.get(staff, (start, start))
            arg = S.text_arg(staff)
            basis = self.any_basis(staff)
            informational = FLAT in basis or PIECE_RATE in basis
            if PIECE_RATE in basis:
                S.write_row(ws, self.style, row, [
                    staff, None, None, None, None, None, None, "—",
                    "piece-rate — no attendance is tracked, so there is nothing to rate",
                ])
                row += 1
                continue
            S.write_row(ws, self.style, row, [
                staff,
                f'=IFERROR(AVERAGEIF($J${first}:$J${last},"<>{NOT_EMPLOYED}",'
                f'$G${first}:$G${last}),"")',
                f'=IFERROR(AVERAGEIFS($I${first}:$I${last},$J${first}:$J${last},'
                f'"<>{NO_DATA}",$J${first}:$J${last},"<>{NOT_EMPLOYED}"),"")',
                f'=COUNTIF($J${first}:$J${last},"Satisfactory")',
                f'=COUNTIF($J${first}:$J${last},"Average")',
                f'=COUNTIF($J${first}:$J${last},"Below Average")',
                f'=COUNTIF($J${first}:$J${last},"{NO_DATA}")',
                f'=IF(C{row}="","{NO_DATA}",IF(C{row}>=$B${bands_row},"Satisfactory",'
                f'IF(C{row}>=$B${bands_row + 1},"Average","Below Average")))',
                ("flat pay — this rating is for information only, it does not affect "
                 "what this person is paid") if informational else "",
            ], [None, S.QTY, S.PERCENT, S.INT, S.INT, S.INT, S.INT, None, None])
            row += 1

    def _monthly_row_for(self, staff, month):
        """Which Monthly Summary row holds this staff-month."""
        m0, _ = self.monthly_rows
        r = m0
        for s in self.staff_order():
            if PIECE_RATE in self.any_basis(s):
                if s == staff:
                    return None
                r += 2
                continue
            for mth in self.months:
                if s == staff and mth.key == month.key:
                    return r
                r += 1
            r += 1
        return None

    # =======================================================================

    def needs_review(self):
        ws = self.sheet(K_REVIEW)
        S.prepare(ws, self.style)
        row = S.title(ws, self.style, 1, f"Needs Review — FY{self.fy}",
                      "Nothing is ever dropped. Anything the engine could not settle "
                      "on its own lands here with the reason, for a person to decide.")
        row = S.header(ws, self.style, row,
                       ["Where", "What", "Why it needs a person", "Suggestion"],
                       [30, 30, 62, 40])
        start = row
        items = list(self.inp.review or [])
        if self.inp.karigar is not None:
            items += list(self.inp.karigar.review or [])
        if self.inp.master is not None:
            for r in self.inp.master.review:
                items.append({"where": r.where, "what": r.what, "reason": r.reason,
                              "detail": r.detail})
        for it in items:
            if not isinstance(it, dict):
                it = {"what": str(it), "reason": ""}
            suggestion = it.get("proposed_merges") or it.get("detail") or ""
            row = S.write_row(ws, self.style, row, [
                str(it.get("where", "")), str(it.get("what", "")),
                str(it.get("reason", "")),
                str(suggestion) if suggestion else "",
            ])
        for n in self.out.notes:
            row = S.write_row(ws, self.style, row, ["workbook", "sheet", n, ""])
        if row == start:
            S.note(ws, self.style, row, "Nothing needed a person this run.")

    # =======================================================================

    def run(self, path) -> Built:
        for name in SHEET_ORDER:
            self.wb.create_sheet(name)
        self.karigar_sheets()
        self.staff_sheets()
        self.read_me_karigar()
        self.read_me_staff()
        self.overview()
        self.needs_review()
        self.wb.save(path)
        self.out.path = str(path)
        self.out.sheets = [ws.title for ws in self.wb]
        return self.out


def build(path, inputs: Inputs) -> Built:
    """Write one financial year's workbook. Returns what was written and what
    each checked cell must recalculate to."""
    return _Build(inputs).run(path)
