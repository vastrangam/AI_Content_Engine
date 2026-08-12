"""Workbooks in, plain rows out.

This is the only module that knows what a spreadsheet is. Everything else works
on lists of lists, so the rules can be tested without a workbook anywhere near
them — and so a CSV, a Google Sheet export or a database table can feed the same
engine later without touching a single rule.

openpyxl is imported lazily, inside the functions, for the same reason: the
engine and its self-tests run without it installed.
"""

from __future__ import annotations

from pathlib import Path


class WorkbookError(RuntimeError):
    pass


def _load(path, data_only=True):
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - depends on the machine
        raise WorkbookError(
            "openpyxl is needed to read spreadsheets.  pip install openpyxl"
        ) from exc
    p = Path(path)
    if not p.exists():
        raise WorkbookError(f"no such workbook: {p}")
    # data_only reads the last value Excel calculated. A workbook that has never
    # been opened in Excel has no cached values, which is why sheet_rows warns.
    return openpyxl.load_workbook(p, data_only=data_only, read_only=True)


def sheet_names(path) -> list[str]:
    wb = _load(path)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def sheet_rows(path, sheet=None) -> list[list]:
    """One sheet as a list of rows, trailing empties trimmed."""
    wb = _load(path)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows = []
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            while values and values[-1] is None:
                values.pop()
            rows.append(values)
        while rows and not rows[-1]:
            rows.pop()
        return rows
    finally:
        wb.close()


def all_sheets(path) -> dict[str, list[list]]:
    """Every sheet, in the order the workbook holds them."""
    wb = _load(path)
    try:
        out = {}
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                values = list(row)
                while values and values[-1] is None:
                    values.pop()
                rows.append(values)
            while rows and not rows[-1]:
                rows.pop()
            out[name] = rows
        return out
    finally:
        wb.close()


def formulas(path) -> list[str]:
    """Every formula string in the workbook — what gate 7 inspects."""
    wb = _load(path, data_only=False)
    try:
        found = []
        for name in wb.sheetnames:
            for row in wb[name].iter_rows(values_only=True):
                for value in row:
                    if isinstance(value, str) and value.startswith("="):
                        found.append(value)
        return found
    finally:
        wb.close()


def cached_values_present(path) -> bool:
    """False when the file has formulas but no calculated values behind them.

    A workbook written by a script and never opened in Excel reads as all-None
    through data_only. Better to say so than to report a payroll of zero.
    """
    wb = _load(path, data_only=False)
    try:
        has_formula = any(
            isinstance(v, str) and v.startswith("=")
            for name in wb.sheetnames
            for row in wb[name].iter_rows(values_only=True)
            for v in row
        )
    finally:
        wb.close()
    if not has_formula:
        return True
    wb = _load(path, data_only=True)
    try:
        return any(
            v is not None
            for name in wb.sheetnames
            for row in wb[name].iter_rows(values_only=True)
            for v in row
        )
    finally:
        wb.close()
